import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from app.database import get_db
from app.services.deal_service import DealService

router = APIRouter(prefix="/reports", tags=["reports"])


@router.get("/pipeline.xlsx")
async def export_pipeline_excel(db=Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    deal_svc = DealService(db)
    deals, _ = await deal_svc.list_deals(limit=500)

    wb = Workbook()

    # Sheet 1: All deals
    ws1 = wb.active
    ws1.title = "Pipeline"
    headers = ["Prospecto", "Línea", "Servicio", "Etapa", "Valor ($)", "Valor Pond. ($)",
               "Responsable", "Próxima Acción", "Fecha Cierre", "Creado"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A7A4A")
        cell.alignment = Alignment(horizontal="center")

    for deal in deals:
        d = deal.model_dump(by_alias=True)
        ws1.append([
            d.get("prospectId", ""),
            d.get("line", ""),
            d.get("serviceType", ""),
            d.get("stage", ""),
            d.get("value") or 0,
            d.get("ponderatedValue") or 0,
            d.get("assignedTo", ""),
            d.get("nextAction", ""),
            str(d.get("expectedCloseDate", "") or ""),
            str(d.get("createdAt", ""))[:10] if d.get("createdAt") else "",
        ])

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 2: By stage
    ws2 = wb.create_sheet("Por Etapa")
    stages: dict[str, dict] = {}
    for deal in deals:
        s = deal.stage
        if s not in stages:
            stages[s] = {"count": 0, "value": 0}
        stages[s]["count"] += 1
        stages[s]["value"] += deal.value or 0

    ws2.append(["Etapa", "Cantidad", "Valor Total ($)"])
    for h in ["A1", "B1", "C1"]:
        ws2[h].font = Font(bold=True)
    for stage, data in stages.items():
        ws2.append([stage, data["count"], data["value"]])

    # Sheet 3: By service
    ws3 = wb.create_sheet("Por Servicio")
    services: dict[str, dict] = {}
    for deal in deals:
        svc = deal.serviceType or "Sin clasificar"
        if svc not in services:
            services[svc] = {"count": 0, "value": 0}
        services[svc]["count"] += 1
        services[svc]["value"] += deal.value or 0

    ws3.append(["Servicio", "Deals", "Valor Total ($)"])
    for h in ["A1", "B1", "C1"]:
        ws3[h].font = Font(bold=True)
    for svc, data in services.items():
        ws3.append([svc, data["count"], data["value"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pipeline_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/quarterly.pdf")
async def export_quarterly_pdf(
    quarter: int = Query(2, ge=1, le=4),
    year: int = Query(2026),
    db=Depends(get_db),
):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER

    deal_svc = DealService(db)
    deals, _ = await deal_svc.list_deals(limit=500)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter,
                            rightMargin=inch * 0.75, leftMargin=inch * 0.75,
                            topMargin=inch, bottomMargin=inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                 textColor=colors.HexColor("#1A7A4A"),
                                 fontSize=22, alignment=TA_CENTER)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                               textColor=colors.HexColor("#1A7A4A"), fontSize=14)
    body_style = styles["Normal"]

    story = []
    story.append(Paragraph("CoimpactoB — Reporte Trimestral", title_style))
    story.append(Paragraph(f"Q{quarter} {year}", title_style))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", body_style))
    story.append(Spacer(1, 0.4 * inch))

    won_deals = [d for d in deals if d.stage == "GANADO"]
    story.append(Paragraph("Deals Ganados en el Trimestre", h2_style))
    story.append(Spacer(1, 0.1 * inch))

    if won_deals:
        table_data = [["Prospecto", "Línea", "Valor ($)", "Responsable"]]
        total_won = 0
        for d in won_deals:
            table_data.append([
                d.prospectId or "",
                d.line or "",
                f"${(d.value or 0):,.0f}",
                d.assignedTo or "",
            ])
            total_won += d.value or 0
        table_data.append(["TOTAL", "", f"${total_won:,.0f}", ""])

        t = Table(table_data, colWidths=[2.2 * inch, 1.5 * inch, 1.3 * inch, 1.5 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A7A4A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F5E9")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Sin deals ganados en el trimestre.", body_style))

    story.append(Spacer(1, 0.4 * inch))
    story.append(Paragraph("Pipeline Activo por Línea de Negocio", h2_style))
    story.append(Spacer(1, 0.1 * inch))

    lines: dict[str, dict] = {}
    for deal in deals:
        if deal.stage in ("GANADO", "PERDIDO"):
            continue
        line = deal.line or "Sin clasificar"
        if line not in lines:
            lines[line] = {"count": 0, "value": 0}
        lines[line]["count"] += 1
        lines[line]["value"] += deal.value or 0

    if lines:
        line_data = [["Línea", "Deals Activos", "Valor Pipeline ($)"]]
        for line, data in lines.items():
            line_data.append([line, str(data["count"]), f"${data['value']:,.0f}"])
        lt = Table(line_data, colWidths=[2.5 * inch, 1.5 * inch, 2.5 * inch])
        lt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A7A4A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F8F4")]),
        ]))
        story.append(lt)

    doc.build(story)
    output.seek(0)

    filename = f"reporte_Q{quarter}_{year}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
